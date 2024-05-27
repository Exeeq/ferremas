from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from .forms import CustomAuthenticationForm
from django.shortcuts import render
from django.db import connection
from django.urls import reverse
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
import requests
from django.http import JsonResponse
from django.db import connection, DataError, IntegrityError
from openpyxl import Workbook
from django.http import HttpResponse
import pytz
from django.utils import timezone
from datetime import datetime
from rest_framework import viewsets
from .serializers import *
from django.core.paginator import Paginator
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl.utils
from django.contrib.auth import views as auth_views

# SERIALIZERS:
class RolUsuarioViewSet(viewsets.ModelViewSet):
    queryset = rolUsuario.objects.all()
    serializer_class = RolUsuarioSerializer

class RegionViewSet(viewsets.ModelViewSet):
    queryset = region.objects.all()
    serializer_class = RegionSerializer

class ComunaViewSet(viewsets.ModelViewSet):
    queryset = comuna.objects.all()
    serializer_class = ComunaSerializer

class UsuarioCustomViewSet(viewsets.ModelViewSet):
    queryset = usuarioCustom.objects.all()
    serializer_class = UsuarioCustomSerializer

class MarcaViewSet(viewsets.ModelViewSet):
    queryset = marca.objects.all()
    serializer_class = MarcaSerializer

class CategoriaProductoViewSet(viewsets.ModelViewSet):
    queryset = categoriaProducto.objects.all()
    serializer_class = CategoriaProductoSerializer

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = producto.objects.all()
    serializer_class = ProductoSerializer

class SucursalViewSet(viewsets.ModelViewSet):
    queryset = sucursal.objects.all()
    serializer_class = SucursalSerializer

class CarritoViewSet(viewsets.ModelViewSet):
    queryset = Carrito.objects.all()
    serializer_class = CarritoSerializer

class ItemCarritoViewSet(viewsets.ModelViewSet):
    queryset = ItemCarrito.objects.all()
    serializer_class = ItemCarritoSerializer

class SeguimientoViewSet(viewsets.ModelViewSet):
    queryset = Seguimiento.objects.all()
    serializer_class = SeguimientoSerializer

class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer

class ItemPedidoViewSet(viewsets.ModelViewSet):
    queryset = ItemPedido.objects.all()
    serializer_class = ItemPedidoSerializer


# VIEWS
def index(request):
	return render(request, 'core/index.html')

@login_required
def shop(request):
    
    productos = lista_productos()
    paginator = Paginator(productos, 4)

    page = request.GET.get('page', 1)
    productos = paginator.get_page(page)
    
    data = {
        'listado': productos,
        'MEDIA_URL': settings.MEDIA_URL,
        'paginator': paginator
    }
    return render(request, 'core/shop.html', data)

def about(request):
	return render(request, 'core/about.html')

def services(request):
	return render(request, 'core/services.html')

@login_required
def contact(request):
	return render(request, 'core/contact.html')

@login_required
def cart(request):
    try:
        usuario = request.user
        carrito = Carrito.objects.get(usuario=usuario)
        items = carrito.itemcarrito_set.all()
        
        api_mindicador = requests.get('https://mindicador.cl/api/')
        divisas = api_mindicador.json()
        tasa_dolar = divisas['dolar']['valor']

        subtotal = sum(item.precio_total() for item in items)
        subtotal_dolar = round(subtotal / tasa_dolar, 2)
        total = subtotal 
        total_dolar = subtotal_dolar

        data = {
            'carrito': carrito,
            'items': items,  
            'subtotal': subtotal,
            'total': total,
            'subtotal_dolar': subtotal_dolar,
            'total_dolar': total_dolar,
            'MEDIA_URL': settings.MEDIA_URL,
        }
        
        return render(request, 'core/cart.html', data)
    
    except Carrito.DoesNotExist:

        data = {
            'carrito': None,
            'items': None,
            'subtotal': 0,
            'total': 0,
            'subtotal_dolar': 0,
            'total_dolar': 0,
            'MEDIA_URL': settings.MEDIA_URL,
        }
        # Muestra un mensaje informativo
        messages.info(request, 'Tu carrito está vacío. Añade productos para continuar.')
        return render(request, 'core/cart.html', data)


@login_required
def checkout(request):
    try:
        regiones = region.objects.all()  
        comunas = comuna.objects.all()
        usuario = request.user
        carrito = Carrito.objects.get(usuario=usuario)
        items = carrito.itemcarrito_set.all()
        
        sucursales = sucursal.objects.all()

        api_mindicador = requests.get('https://mindicador.cl/api/')
        divisas = api_mindicador.json()
        tasa_dolar = divisas['dolar']['valor']

        subtotal = sum(item.precio_total() for item in items)
        subtotal_dolar = round(subtotal / tasa_dolar, 2)
        total = subtotal 
        total_dolar = subtotal_dolar

        subtotal_dolar_str = '{:.2f}'.format(subtotal_dolar) 
        total_dolar_str = '{:.2f}'.format(total_dolar)

        data = {
            'carrito': carrito,
            'items': items,  
            'subtotal': subtotal,
            'total': total,
            'subtotal_dolar': subtotal_dolar_str,
            'total_dolar': total_dolar_str,
            'regiones': regiones,
            'comunas': comunas,
            'sucursales': sucursales,
            'MEDIA_URL': settings.MEDIA_URL,
        }
        
        return render(request, 'core/checkout.html', data)
    
    except Exception as e:
        print("ERROR EN CHECKOUT: ", e)


@login_required
def thankyou(request):
	return render(request, 'core/thankyou.html')

def register(request):
    form = RegisterForm()
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            # Asignar el rol por defecto Cliente
            user.idRol = rolUsuario.objects.get(nombreRol='Cliente')
            user.idComuna = form.cleaned_data['comuna']
            
            # Guardar el correo_usuario en el campo email predeterminado
            user.email = user.correo_usuario

            user.save()
            return redirect("index")

    return render(request, 'registration/register.html', {'form': form})

class CustomLoginView(LoginView):
    authentication_form = CustomAuthenticationForm
    template_name = 'registration/register.html'
    success_url = reverse_lazy('core/index.html')

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)
        return super().form_valid(form)

def handle_uploaded_file(f):
    with open('media/productos/' + f.name, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

#VISTAS CRUD PRODUCTOS (AÑADIR, ACTUALIZAR, ELIMINAR)
def addProduct(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            nombreProducto = form.cleaned_data['nombreProducto']
            precioProducto = form.cleaned_data['precioProducto']
            stockProducto = form.cleaned_data['stockProducto']
            imagenProducto = request.FILES['imagenProducto']
            descripcionProducto = form.cleaned_data['descripcionProducto']
            idcategoriaProducto = form.cleaned_data['idcategoriaProducto'].pk  
            idMarca = form.cleaned_data['idMarca'].pk  

            with open('media/productos/' + imagenProducto.name, 'wb+') as destination:
                for chunk in imagenProducto.chunks():
                    destination.write(chunk)
                    
            error_msg = agregar_producto(nombreProducto, precioProducto, stockProducto, imagenProducto.name, descripcionProducto, idcategoriaProducto, idMarca)
            if error_msg:
                form.save()
                return render(request, 'core/shop.html', {'form': form, 'error_msg': error_msg})
            else:
                return redirect("shop")

    else:
        form = ProductoForm()
    return render(request, 'core/addproduct.html', {'form': form})


#LISTAR PRODUCTOS:
def lista_productos():
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()

    cursor.callproc("SP_GET_PRODUCTOS", [""])

    lista = []
    for x in cursor:
         lista.append(x)

    return lista

def agregar_producto(nombreProducto, precioProducto, stockProducto, imagenProducto, descripcionProducto, idcategoriaProducto, idMarca):
    try:
        django_cursor = connection.cursor()
        cursor = django_cursor.connection.cursor()

        #Llamar al procedimiento almacenado
        cursor.callproc("SP_POST_PRODUCTO", [nombreProducto, precioProducto, stockProducto, imagenProducto, descripcionProducto, idcategoriaProducto, idMarca])
        return None 
    
    except Exception as e:
        error_msg = "Hubo un error al agregar el producto. Por favor, inténtalo de nuevo más tarde."
        print("ERROR EN AGREGAR PRODUCTO: ", e)
        return error_msg
    
def detalle_producto(request, idProducto):
    producto_instance = producto.objects.get(idProducto=idProducto) 
    api_mindicador = requests.get('https://mindicador.cl/api/')
    divisas = api_mindicador.json()
    tasa_dolar = divisas['dolar']['valor']

    precio_dolares = round(producto_instance.precioProducto / tasa_dolar, 2)

    data = {
         'producto':producto_instance,
         'MEDIA_URL': settings.MEDIA_URL,
         'precio_dolares': precio_dolares,
    }
    return render(request, 'core/detalle_producto.html', data)

def modificar_producto(request, idProducto):
    producto_instance = get_object_or_404(producto, idProducto=idProducto)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto_instance)
        if form.is_valid():
            try:
                nombreProducto = form.cleaned_data['nombreProducto']
                precioProducto = form.cleaned_data['precioProducto']
                stockProducto = form.cleaned_data['stockProducto']
                imagenProducto = request.FILES.get('imagenProducto', None)
                descripcionProducto = form.cleaned_data['descripcionProducto']
                idMarca = form.cleaned_data['idMarca'].pk
                idcategoriaProducto = form.cleaned_data['idcategoriaProducto'].pk

                if imagenProducto:
                    with open('media/productos/' + imagenProducto.name, 'wb+') as destination:
                        for chunk in imagenProducto.chunks():
                            destination.write(chunk)
                    imagenProducto_name = imagenProducto.name
                else:
                    imagenProducto_name = producto_instance.imagenProducto

                with connection.cursor() as cursor:
                    cursor.callproc('SP_PUT_PRODUCTO', [
                        idProducto,
                        nombreProducto,
                        precioProducto,
                        stockProducto,
                        imagenProducto_name,
                        descripcionProducto,
                        idMarca,
                        idcategoriaProducto
                    ])
                messages.success(request, 'Producto modificado exitosamente.')
                return redirect('detalle_producto', idProducto=idProducto)

            except DataError as e:
                messages.error(request, "Hubo un problema al ingresar los datos, revise nuevamente por favor.")
                print(e)
            except IntegrityError as e:
                messages.error(request, "Hubo un problema con la integridad de los datos. Por favor, inténtalo de nuevo.")
            except Exception as e:
                messages.error(request, "Ocurrió un error inesperado. Por favor, inténtalo de nuevo.")

    else:
        form = ProductoForm(instance=producto_instance)
    return render(request, 'core/modificar_producto.html', {'form': form})

def eliminar_producto(request, idProducto):
    with connection.cursor() as cursor:
        cursor.callproc('SP_DELETE_PRODUCTO', [idProducto])
    return redirect(to='shop')


#PANEL DE ADMINISTRACIÓN (ROL ADMINISTRADOR):
def panel_administracion(request):
    return render(request, 'core/panel_administracion.html') 

def gestion_usuarios(request):
    data = {
        'usuarios': lista_usuarios()
    }
    return render(request, 'core/gestion_usuarios.html', data)


#GESTIÓN DE USUARIOS
def lista_usuarios():
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()

    cursor.callproc("SP_GET_USUARIOS", [""])
    lista = []
    for fila in cursor:
         lista.append(fila)
    return lista

def eliminar_usuario(request, id):
    try:
        django_cursor = connection.cursor()
        cursor = django_cursor.connection.cursor()

        cursor.callproc("SP_DELETE_USUARIO", [id])
        django_cursor.connection.commit()
        return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})
        
    except Exception as e:
        print("Error en eliminar usuario: ", e)
        return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})


def modificar_usuario(request, p_id):
    usuario = usuarioCustom.objects.get(id=p_id)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            p_username = form.cleaned_data['username'] 
            p_run = form.cleaned_data['run'] 
            p_pnombre = form.cleaned_data['pnombre'] 
            p_ap_paterno = form.cleaned_data['ap_paterno'] 
            p_correo_usuario = form.cleaned_data['correo_usuario'] 
            p_fecha_nacimiento = form.cleaned_data['fecha_nacimiento'] 
            p_direccion = form.cleaned_data['direccion']
            p_idComuna = form.cleaned_data['idComuna'].pk
            p_idRol = form.cleaned_data['idRol'].pk

            #Llamar al procedimiento almacenado para actualizar al usuario
            with connection.cursor() as cursor:
                cursor.callproc("SP_PUT_USUARIO", [
                     p_id,
                     p_username,
                     p_run,
                     p_pnombre,
                     p_ap_paterno,
                     p_correo_usuario,
                     p_fecha_nacimiento,
                     p_direccion,
                     p_idComuna,
                     p_idRol
                ])
            messages.success(request, 'Usuario modificado correctamente!')
            form.save()
            return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, 'core/modificar_usuario.html', {'form': form})

def post_usuario(p_username, p_run, p_pnombre, p_ap_paterno, p_correo_usuario, p_fecha_nacimiento, p_direccion, p_idComuna, p_idRol, p_password):
    try:
        django_cursor = connection.cursor()
        cursor = django_cursor.connection.cursor()

        p_password_encrypted = make_password(p_password)
        cursor.callproc("SP_POST_USUARIO", [
             p_username,
             p_run,
             p_pnombre,
             p_ap_paterno,
             p_correo_usuario,
             p_fecha_nacimiento,
             p_direccion,
             p_idComuna,
             p_idRol,
             p_password_encrypted
        ])
        usuarioCustom.objects.create_user(username=p_username, email=p_correo_usuario, password=p_password)
        return print("Funciono post usuario!")
    except Exception as e:
        return print("ERROR EN POST USUARIO: ", e)
    

def add_user(request):
    if request.method == 'POST':
        form = RegisterUserAdminForm(request.POST)
        if form.is_valid():
            form.save()
            p_username = form.cleaned_data['username']
            p_run = form.cleaned_data['run']
            p_pnombre = form.cleaned_data['pnombre']
            p_ap_paterno = form.cleaned_data['ap_paterno']
            p_correo_usuario = form.cleaned_data['correo_usuario']
            p_fecha_nacimiento = form.cleaned_data['fecha_nacimiento']
            p_direccion = form.cleaned_data['direccion']
            p_idComuna = form.cleaned_data['idComuna'].pk
            p_idRol = form.cleaned_data['idRol'].pk
            p_password = form.cleaned_data['password1']  

            post_usuario(p_username,
                         p_run,
                         p_pnombre,
                         p_ap_paterno,
                         p_correo_usuario,
                         p_fecha_nacimiento,
                         p_direccion,
                         p_idComuna,
                         p_idRol,
                         p_password)
            messages.success(request, 'Usuario agregado correctamente!')
            return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})
    else:
        form = RegisterUserAdminForm()
    return render(request, 'core/addusuario.html', {'form': form})
            
            
#CARRITO DE COMPRAS:
def agregar_al_carrito(request, idProducto):
    producto_cart = get_object_or_404(producto, pk=idProducto)
    carrito, created = Carrito.objects.get_or_create(usuario=request.user)
    item, item_created = ItemCarrito.objects.get_or_create(carrito=carrito, producto=producto_cart)

    if not item_created:
        item.cantidad += 1
        item.save()
    else:
        item.cantidad = 1
        item.save()

    # Disminuir el stock del producto
    producto_cart.disminuir_stock(1)

    return redirect(to="cart")

def eliminar_del_carrito(request, itemcarrito_id):
    item = get_object_or_404(ItemCarrito, pk=itemcarrito_id, carrito__usuario=request.user)
    producto = item.producto

    # Incrementar la cantidad disponible en el stock del producto
    producto.stockProducto += item.cantidad
    producto.save()

    item.delete()

    return redirect('cart')


#VISTAS RELACIONADAS A LA CREACIÓN DEL PEDIDO, BOLETA, ETC:
def crear_pedido(request):
    if request.method == 'POST':
        usuario = request.user
        carrito = usuario.carrito

        # Recibir los datos del formulario
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        run = request.POST.get('run') 
        tipo_entrega = request.POST.get('tipo_entrega')
        sucursal_id = request.POST.get('sucursal')

        print(sucursal_id)
        # Buscar la instancia de la sucursal utilizando el ID
        if sucursal_id:
            sucursal_instance = sucursal.objects.get(idSucursal=sucursal_id)
        else:
            sucursal_instance = None

        if tipo_entrega == "retiro_tienda":
            pedido = Pedido.objects.create(
                carrito=carrito,
                numero=str(uuid.uuid4()),
                nombre=nombre,
                apellido=apellido,
                direccion=None,
                region=None,
                comuna=None,
                correo=None,
                run=run,
                sucursal=sucursal_instance,
                tipo_entrega=tipo_entrega,
            )
        else:
            # Recibir los datos adicionales para envío a domicilio
            direccion = request.POST.get('direccion', '')
            region_id = request.POST.get('region')
            comuna_id = request.POST.get('comuna')
            correo = request.POST.get('correo')

            # Verificar si region_id es una cadena vacía
            if region_id:
                region_obj = region.objects.get(idRegion=region_id)
            else:
                region_obj = None

            # Verificar si comuna_id es una cadena vacía
            if comuna_id:
                comuna_obj = comuna.objects.get(idComuna=comuna_id)
            else:
                comuna_obj = None

            pedido = Pedido.objects.create(
                carrito=carrito,
                numero=str(uuid.uuid4()),
                nombre=nombre,
                apellido=apellido,
                direccion=direccion,
                region=region_obj,
                comuna=comuna_obj,
                correo=correo,
                run=None,
                sucursal=sucursal_instance,
                tipo_entrega=tipo_entrega,
            )

        items_carrito = carrito.itemcarrito_set.all()
        for item in items_carrito:
            ItemPedido.objects.create(pedido=pedido, producto=item.producto, cantidad=item.cantidad)

        carrito.productos.clear()

        return JsonResponse({'success': True, 'numero_pedido': pedido.numero})
    else:
        return JsonResponse({'success': False})
    
def boleta(request, numero_pedido):
    usuario = request.user
    pedido = Pedido.objects.get(numero = numero_pedido)
    pedidos = Pedido.objects.filter(carrito__usuario=usuario)

    data = {
         'pedido': pedido,
         'pedidos': pedidos,
         'MEDIA_URL': settings.MEDIA_URL,

    }

    return render(request, 'core/boleta.html', data)

@login_required
def mis_pedidos(request):
    usuario = request.user
    pedidos = Pedido.objects.filter(carrito__usuario=usuario)

    data = {
        'pedidos': pedidos,
        'MEDIA_URL': settings.MEDIA_URL,
    }
    return render(request, 'core/mis_pedidos.html', data)

@login_required
def administrar_pedidos(request):
    pedidos = Pedido.objects.all()
    form = EstadoPedido()

    if request.method == 'POST':
        form = EstadoPedido(request.POST)
        if form.is_valid():
            pedido_numero = form.cleaned_data['pedido_numero']
            estado = form.cleaned_data['estado']
            pedido = Pedido.objects.get(numero=pedido_numero)
            pedido.estado = estado
            pedido.save()
    
    data = {
        'pedidos':pedidos,
        'form': form
    }
    return render(request, 'core/administrar_pedidos.html', data)

def cambiar_estado(request, numero_orden):
    pedido = Pedido.objects.get(numero=numero_orden)
    form = EstadoPedido(request.POST or None, initial={'estado': pedido.estado.descripcion})

    if request.method == 'POST' and form.is_valid():
        estado_descripcion = form.cleaned_data['estado']
        seguimiento, created = Seguimiento.objects.get_or_create(descripcion=estado_descripcion)
        pedido.estado = seguimiento
        pedido.save()
        # Realizar cualquier acción adicional después de actualizar el estado
        return redirect(reverse('administrar_pedidos'))  # Redirige a donde quieras

    data = {
        'pedidos': Pedido.objects.all(),
        'form': form,
    }

    return render(request, 'core/administrar_pedidos.html', data)

#SOPORTE CONTACTO ADMINISTRADOR:
def soporte_contacto(request):
    return render(request, 'core/soporte_contacto.html')

def pedidos_entregados(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    pedidos = []

    # Convertir los valores de mes y anio a enteros si están presentes, de lo contrario a 0
    mes_num = int(mes) if mes and mes.isdigit() else 0
    anio_num = int(anio) if anio and anio.isdigit() else 0

    with connection.cursor() as cursor:
        cursor.callproc('FiltrarPedidosEntregados', [mes_num, anio_num])
        resultados = cursor.fetchall()
        for row in resultados:
            pedido = {
                'id': row[0],  # ID del pedido
                'numero': row[1],  # Número de pedido
                'fecha': row[2],  # Fecha del pedido
                'carrito_id': row[3],  # ID del carrito
                'estado': row[4],  # Estado
                'apellido': row[5],  # Apellido del cliente
                'comuna': row[6],  # Nombre de la comuna
                'correo': row[7],  # Correo electrónico del cliente
                'direccion': row[8],  # Dirección del cliente
                'nombre': row[9],  # Nombre del cliente
                'region': row[10],  # Nombre de la región
                'sucursal_id': row[11],  # ID de la sucursal
                'run': row[12],  # RUN del cliente
                'tipo_entrega': row[13],  # Tipo de entrega
                'total_pagado': row[14]  # Total pagado
            }
            pedidos.append(pedido)

    meses = [str(i).zfill(2) for i in range(1, 13)]
    current_year = datetime.now().year
    years = list(range(2020, current_year + 1))
    
    data = {
        'pedidos': pedidos,
        'meses': meses,
        'years': years,
        'selected_mes': mes,
        'selected_anio': anio,
    }
    return render(request, 'core/pedidos_entregados.html', data)

#GENERAR INFORMES:
def generar_informes(request):
    # Obtener parámetros de filtro
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    # Convertir los valores de mes y anio a enteros si están presentes, de lo contrario a 0
    mes_num = int(mes) if mes and mes.isdigit() else 0
    anio_num = int(anio) if anio and anio.isdigit() else 0

    with connection.cursor() as cursor:
        cursor.callproc('FiltrarPedidosEntregados', [mes_num, anio_num])
        resultados = cursor.fetchall()

        # Crear un libro de trabajo y una hoja de trabajo
        wb = Workbook()
        ws = wb.active

        # Definir encabezados de columna
        headers = [
            'Número de Pedido', 'Fecha', 'Cliente', 'Dirección', 
            'Comuna', 'Región', 'Estado Actual', 'Total Pagado'
        ]
        
        # Aplicar estilos a los encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        ws.append(headers)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # Iterar sobre los pedidos filtrados y añadirlos a la hoja de trabajo
        for row in resultados:
            pedido = [
                row[1],  # Número de Pedido
                row[2].strftime('%d/%m/%Y'),  # Fecha
                f"{row[9]} {row[5]}",  # Cliente (Nombre Apellido)
                row[8],  # Dirección
                row[6],  # Comuna
                row[10],  # Región
                row[4],  # Estado Actual
                f"${row[14]:,.2f} CLP"  # Total Pagado
            ]
            ws.append(pedido)

        # Ajustar el ancho de las columnas
        column_widths = [20, 15, 30, 30, 20, 20, 15, 20]
        for i, col_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_width

        # Obtener el nombre del archivo basado en los filtros aplicados
        filename = f'ventas_entregadas_{mes or "None"}_{anio or "None"}.xlsx'

        # Crear una respuesta HTTP con el contenido del libro de trabajo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        
        return response

def obtener_sucursales(request):
    try:
        sucursales = sucursal.objects.all()
        data = [{'idSucursal': s.idSucursal, 'nombreSucursal': s.nombreSucursal, 'direccionSucursal': s.direccionSucursal} for s in sucursales]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

#Perfil
@login_required
def perfil_usuario(request):
    user = request.user
    if request.method == 'POST':
        form = UsuarioCustomForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente!')
            return redirect('perfil_usuario')
    else:
        form = UsuarioCustomForm(instance=user)
    
    return render(request, 'core/perfil_usuario.html', {'form': form})

#Correo
class CustomPasswordResetView(auth_views.PasswordResetView):
    email_template_name = 'registration/password_reset_email.html'

class PasswordResetView(auth_views.PasswordResetView):
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'

